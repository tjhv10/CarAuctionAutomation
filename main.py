import os
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
import openpyxl
from openpyxl import load_workbook
excel_file = "car_data.xlsx"
if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Car Data"
    headers = ["Loyer" ,"Link", "Name", "Discription", "Hand","Picture 1","Picture 2","Picture 3"]
    ws.append(headers)

url = 'https://cars.bidspirit.com/ui/home?lang=he'

web = webdriver.Chrome()
web.get(url)
sleep(3)
web.find_element(By.XPATH, '//*[@id="onetrust-reject-all-handler"]').click()

for i in range(18, 21):
    print(i)
    sleep(5)
    web.find_element(By.XPATH, '//*[@id="mainView"]/div/div/div[2]/div/div/div[2]/div/div[5]/div[1]/div[2]/div[2]/div/div[' + str(i) + ']/div/div').click()
    sleep(3)
    try:
        num_of_cars = int(web.find_element(By.XPATH, '//*[@id="mainView"]/div/div/div[2]/div/div/div/div/div[1]/div[2]/div[1]/div[1]/div').text.split(' ')[0])
    except NoSuchElementException:
        num_of_cars = 0
    j = 1
    for k in range(1, num_of_cars + 1):
        print(k)
        try:
            loyer = web.find_element(By.XPATH, '//*[@id="mainView"]/div/div/div[2]/div/div/div/div/div[1]/div[1]/div[1]/div[2]/h2/a').text
        except:
            loyer = "None"
        try:
            sleep(3)
            web.find_element(By.XPATH, '//*[@id="mainView"]/div/div/div[2]/div/div/div/div/div[2]/div[1]/div/div[1]/div['+str(j)+']/div[1]').click()
        except NoSuchElementException:
            sleep(3)
            # todo find how to find the fowerd button
            try:
                web.find_element(By.XPATH, '//*[@id="mainView"]/div/div/div[2]/div/div/div/div/div[2]/div[2]/div/a[6]').click()
            except:
                try:
                    web.find_element(By.XPATH, '//*[@id="mainView"]/div/div/div[2]/div/div/div/div/div[2]/div[2]/div/a[7]').click()
                except:
                    try:
                        web.find_element(By.XPATH, '//*[@id="mainView"]/div/div/div[2]/div/div/div/div/div[2]/div[2]/div/a[5]').click()
                    except:
                        web.find_element(By.XPATH, '//*[@id="mainView"]/div/div/div[2]/div/div/div/div/div[1]/div[3]/div[1]/div/a[5]').click()
                    
            sleep(3)
            j = 1
            web.find_element(By.XPATH, '//*[@id="mainView"]/div/div/div[2]/div/div/div/div/div[2]/div[1]/div/div[1]/div['+str(j)+']/div[1]/div').click()
        j += 1
        sleep(5)
        l = 1
        try:
            web.find_element(By.XPATH,'//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[1]/div[8]/div[1]/div/img').click()
            sleep(3)
            pic1 = web.find_element(By.XPATH,'//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[2]/div/div/div[2]/div/div[2]/div/img').get_attribute('src')
            print(pic1)
            sleep(2)
            web.find_element(By.XPATH,'//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[2]/div/div/div[1]/div[3]').click()
            sleep(2)
            pic2 = web.find_element(By.XPATH,'//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[2]/div/div/div[2]/div/div[2]/div/img').get_attribute('src')
            sleep(2)
            web.find_element(By.XPATH,'//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[2]/div/div/div[1]/div[3]').click()
            sleep(2)
            pic3 = web.find_element(By.XPATH,'//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[2]/div/div/div[2]/div/div[2]/div/img').get_attribute('src')
            sleep(2)
            
        except:
            pic1 = "None"
            pic2 = "None"
            pic3 = "None"
        web.back()
        sleep(3)
        try:
            while web.find_element(By.XPATH,'//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[1]/div[11]/table/tbody/tr['+str(l)+']/th').text != "יד":
                l+=1
            hand = web.find_element(By.XPATH, '//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[1]/div[11]/table/tbody/tr['+str(l)+']/td').text
        except:
            hand = "None"
            pass
        try:
            link = web.current_url
        except:
            link = "None"
        try:
            name = web.find_element(By.XPATH, '//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[1]/div[5]/h1').text
        except:
            name = "None"
        try:
            disc = web.find_element(By.XPATH, '//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[1]/div[5]/h2').text
        except:
            disc = "None"
        ws.append([loyer,link, name, disc, hand,pic1,pic2,pic3])    
        wb.save(excel_file)
        web.back()
    
    if num_of_cars == 0:
        try: 
            loyer = web.find_element(By.XPATH, '//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[1]/div[1]/div[1]/div[2]/div/a').text
        except:
            loyer = "None"
        try:
            web.find_element(By.XPATH,'//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[1]/div[8]/div[1]/div/img').click()
            sleep(3)
            pic1 = web.find_element(By.XPATH,'//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[2]/div/div/div[2]/div/div[2]/div/img').get_attribute('src')
            sleep(2)
            web.find_element(By.XPATH,'//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[2]/div/div/div[1]/div[3]').click()
            sleep(2)
            pic2 = web.find_element(By.XPATH,'//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[2]/div/div/div[2]/div/div[2]/div/img').get_attribute('src')
            sleep(2)
            web.find_element(By.XPATH,'//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[2]/div/div/div[1]/div[3]').click()
            sleep(2)
            pic3 = web.find_element(By.XPATH,'//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[2]/div/div/div[2]/div/div[2]/div/img').get_attribute('src')
            sleep(2)
            web.back()
            sleep(3)
            l=1
            while web.find_element(By.XPATH,'//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[1]/div[11]/table/tbody/tr['+str(l)+']/th').text != "יד":
                l+=1
            hand = web.find_element(By.XPATH, '//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[1]/div[11]/table/tbody/tr['+str(l)+']/td').text
        except:
            pic1 = "None"
            pic2 = "None"
            pic3 = "None"
            hand = "None"
            pass
        link = web.current_url
        try:
            name = web.find_element(By.XPATH, '//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[1]/div[5]/h1').text
            disc = web.find_element(By.XPATH, '//*[@id="mainView"]/div/div/div[2]/div/div/div/div[1]/div/div/div[1]/div[5]/h2').text
            ws.append([loyer ,link, name, disc, hand,pic1,pic2,pic3])
            wb.save(excel_file)
        except NoSuchElementException:
            pass
    web.back()
    sleep(5)

wb.save(excel_file)
web.quit()
